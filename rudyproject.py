import os
import sys
import json
from time import sleep
from datetime import datetime

from models.store import Store
from models.product import Product
from models.variant import Variant
import glob
import requests

import threading
from openpyxl import Workbook
from openpyxl.drawing.image import Image as Imag
from PIL import Image
from lxml import html

import warnings
warnings.filterwarnings("ignore")

class myScrapingThread(threading.Thread):
    def __init__(self, threadID: int, name: str, obj, product_url: str) -> None:
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.product_url = product_url
        self.obj = obj
        self.status = 'in progress'
        pass

    def run(self):
        self.obj.get_product_details(self.product_url)
        self.status = 'completed'

    def active_threads(self):
        return threading.activeCount()

class RudyProject_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str) -> None:
        self.DEBUG = DEBUG
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.thread_list = []
        self.thread_counter = 0
        # self.chrome_options = Options()
        # self.chrome_options.add_argument('--disable-infobars')
        # self.chrome_options.add_argument("--start-maximized")
        # self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        # self.args = ["hide_console", ]
        # # self.browser = webdriver.Chrome(options=self.chrome_options, service_args=self.args)
        # self.browser = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=self.chrome_options)
        self.data = []
        # self.ref_json_data = None
        pass

    def controller(self, store: Store):
        try:
            product_urls = list()
            print("Getting all product URL's from Sunglasses category\n")
            for sub_category in self.get_all_sub_category_urls(store):
                # sub_category_name = sub_category.get('sub_category_name')
                sub_category_url = sub_category.get('sub_category_url')
                product_urls += self.get_all_product_urls_from_category_page(sub_category_url)
                product_urls = list(dict.fromkeys(product_urls))
                
            start_time = datetime.now()
            print(f'Type: Sunglasses | Total products: {len(product_urls)}')
            print(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

            self.print_logs(f'Type: Sunglasses | Total products: {len(product_urls)}')
            self.print_logs(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
            if len(product_urls) > 0:
                self.printProgressBar(0, len(product_urls), prefix = 'Progress:', suffix = 'Complete', length = 50)

            for index, product_url in enumerate(product_urls):
                self.create_thread(product_url)
                if len(self.thread_list) > 20:                   
                    self.wait_for_thread_list_to_complete()
                # self.get_product_details(product_url)
                self.printProgressBar(index + 1, len(product_urls), prefix = 'Progress:', suffix = 'Complete', length = 50)
                
            end_time = datetime.now()
            print(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
            print('Duration: {}\n'.format(end_time - start_time))

            self.print_logs(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
            self.print_logs('Duration: {}\n'.format(end_time - start_time))

        except Exception as e:
            if self.DEBUG: print(f'Exception in scraper controller: {e}')
            else: pass
        finally: 
            self.wait_for_thread_list_to_complete()
            self.save_to_json(self.data)

    def get_headers(self, referer: str) -> dict:
        return {
            'authority': 'safilo.my.site.com',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9',
            'referer': referer,
            'sec-ch-ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        }
    
    def get_all_sub_category_urls(self, store: Store) -> list[dict]:
        sub_categories: list[dict] = list() 
        try:
            response = requests.get(url=store.link, headers=self.get_headers(store.link))
            if response.status_code == 200:
                doc_tree = html.fromstring(response.text)
                for a_tag in doc_tree.xpath('//div[@id="block-views-block-menu-sports-block-eyewear"]')[0].xpath('.//li/a'):
                    sub_category_name = a_tag.xpath('./text()')[0]
                    sub_category_url = 'https://www.rudyproject.com{}'.format(a_tag.xpath('./@href')[0])
                    new_sub_category = { 'sub_category_name': sub_category_name, 'sub_category_url': sub_category_url }
                    if new_sub_category not in sub_categories:
                        sub_categories.append(new_sub_category)
            else: self.print_logs(f'{response.status_code} for {store.link}')
        except Exception as e:
            self.print_logs(f'Exception in get_all_sub_category_urls: {str(e)}')
            if self.DEBUG: print(f'Exception in get_all_sub_category_urls: {str(e)}')
            else: pass 
        finally: return sub_categories

    def get_all_product_urls_from_category_page(self, sub_category_url: str) -> list[str]:
        product_urls: list[str] = list()
        try:
            response = requests.get(url=sub_category_url, headers=self.get_headers(sub_category_url))
            if response.status_code == 200:
                doc_tree = html.fromstring(response.text)
                for a_tag in doc_tree.xpath('//div[@id="block-rp-theme-views-block-variations-catalog-block"]')[0].xpath('.//a[@class="text-black"]'):
                    product_url = 'https://www.rudyproject.com{}'.format(str(a_tag.xpath('./@href')[0]).strip())
                    if product_url not in product_urls:
                        product_urls.append(product_url)
                product_urls += self.get_products_from_next_pages(sub_category_url, response.text)
            else: self.print_logs(f'{response.status_code} for {sub_category_url}')
        except Exception as e:
            self.print_logs(f'Exception in get_all_product_urls_from_category_page: {str(e)}')
            if self.DEBUG: print(f'Exception in get_all_product_urls_from_category_page: {str(e)}')
            else: pass
        finally: 
            product_urls = list(dict.fromkeys(product_urls))
            return product_urls

    def get_products_from_next_pages(self, url: str, response_text: str):
        product_urls: list[str] = list()
        try:
            doc_tree = html.fromstring(response_text)
            while doc_tree.xpath('//a[@rel="next"]'):
                next_page_link = doc_tree.xpath('//a[@rel="next"]/@href')[0]
                next_page_url = '{}{}'.format(url, next_page_link)

                response = requests.get(url=next_page_url, headers=self.get_headers(next_page_url))
                if response.status_code == 200:
                    doc_tree = html.fromstring(response.text)
                    for a_tag in doc_tree.xpath('//div[@id="block-rp-theme-views-block-variations-catalog-block"]')[0].xpath('.//a[@class="text-black"]'):
                        product_url = 'https://www.rudyproject.com{}'.format(a_tag.xpath('./@href')[0])
                        product_urls.append(product_url)

                else: 
                    self.print_logs(f'{response.status_code} for {next_page_url}')
                    break

        except Exception as e:
            self.print_logs(f'Exception in get_products_from_next_pages: {str(e)}')
            if self.DEBUG: print(f'Exception in get_products_from_next_pages: {str(e)}')
            else: pass
        finally: return product_urls

    def get_product_details(self, product_url: str):
        try:
            response = requests.get(url=product_url, headers=self.get_headers(product_url))
            if response.status_code == 200:
                doc_tree = html.fromstring(response.text)

                product = Product()
                product.name = doc_tree.xpath('//div[contains(@class, "variants-item") and contains(@class, "active")]')[0].xpath('.//img/@alt')[0]
                product.brand = 'Rudy Project'
                
                product.url = product_url
                
                image_urls = list()

                for src in doc_tree.xpath(f'//img[@alt="{str(product.name).strip()}" and @class="img-fluid"]/@src'):
                    image_url = f'https://www.rudyproject.com{str(src).strip()}'
                    if image_url not in image_urls:
                        image_urls.append(image_url)
                
                if image_urls: product.image = image_urls[0]
                
                product.metafields.for_who = 'Unisex'
                try: product.metafields.frame_color = doc_tree.xpath('//div[contains(@class, "variation_attribute_frame_color")]/div[contains(@class, "item")]/text()')[0]
                except: pass
                try: product.metafields.lens_color = doc_tree.xpath('//div[contains(@class, "variation_attribute_lens_color")]/div[contains(@class, "item")]/text()')[0]
                except: pass
                try: product.metafields.lens_material = doc_tree.xpath('//div[contains(@class, "variation_attribute_lens")]/div[contains(@class, "item")]/text()')[0]
                except: pass

                variant = Variant()
                try: variant.sku = doc_tree.xpath('//div[contains(@class, "variation_field_sku")]/div[contains(@class, "item")]/text()')[0]
                except: pass
                try: variant.listing_price = str(doc_tree.xpath('//span[contains(@class, "variation_price")]/text()')[0]).strip().replace('€', '')
                except: pass
                product.add_single_variant(variant)

                self.data.append(product)
                self.save_to_json(self.data)

        except Exception as e:
            self.print_logs(f'Exception in get_product_details: {str(e)}')
            if self.DEBUG: print(f'Exception in get_product_details: {str(e)}')
            else: pass

    def create_thread(self, product_url: str):
        thread_name = "Thread-"+str(self.thread_counter)
        self.thread_list.append(myScrapingThread(self.thread_counter, thread_name, self, product_url))
        self.thread_list[self.thread_counter].start()
        self.thread_counter += 1

    def is_thread_list_complted(self):
        for obj in self.thread_list:
            if obj.status == "in progress":
                return False
        return True

    def wait_for_thread_list_to_complete(self):
        while True:
            result = self.is_thread_list_complted()
            if result: 
                self.thread_counter = 0
                self.thread_list.clear()
                break
            else: sleep(1)

    def save_to_json(self, products: list[Product]):
        try:
            json_products = []
            for product in products:
                json_varinats = []
                for index, variant in enumerate(product.variants):
                    json_varinat = {
                        'position': (index + 1), 
                        'title': variant.title, 
                        'sku': variant.sku, 
                        'inventory_quantity': variant.inventory_quantity,
                        'found_status': variant.found_status,
                        'wholesale_price': variant.wholesale_price,
                        'listing_price': variant.listing_price, 
                        'barcode_or_gtin': variant.barcode_or_gtin,
                        'size': variant.size,
                    }
                    json_varinats.append(json_varinat)
                json_product = {
                    'brand': product.brand, 
                    'number': product.number, 
                    'name': product.name, 
                    'frame_code': product.frame_code,  
                    'lens_code': product.lens_code, 
                    'lens_color': product.metafields.lens_color,
                    'frame_color': product.metafields.frame_color,
                    # 'status': product.status, 
                    # 'type': product.type, 
                    'url': product.url, 
                    'metafields': [
                        { 'key': 'for_who', 'value': product.metafields.for_who },
                        # { 'key': 'product_size', 'value': product.metafields.product_size }, 
                        { 'key': 'lens_material', 'value': product.metafields.lens_material }, 
                        { 'key': 'lens_technology', 'value': product.metafields.lens_technology }, 
                        { 'key': 'frame_material', 'value': product.metafields.frame_material }, 
                        { 'key': 'frame_shape', 'value': product.metafields.frame_shape },
                        { 'key': 'gtin1', 'value': product.metafields.gtin1 }, 
                        { 'key': 'img_url', 'value': product.image }
                    ],
                    'variants': json_varinats
                }
                json_products.append(json_product)
            
           
            with open(self.result_filename, 'w') as f: json.dump(json_products, f)
            
        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            else: pass

    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

    def printProgressBar(self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total: 
            print()

def read_data_from_json_file(DEBUG, result_filename: str):
    data = []
    try:
        files = glob.glob(result_filename)
        if files:
            f = open(files[-1])
            json_data = json.loads(f.read())

            for json_d in json_data:
                name = str(json_d['name']).strip()
                
                frame_color = str(json_d['frame_color']).strip().title()
                lens_color = str(json_d['lens_color']).strip().title()
                
                for json_metafiels in json_d['metafields']:
                    if json_metafiels['key'] == 'img_url':img_url = str(json_metafiels['value']).strip()
                for json_variant in json_d['variants']:
                    sku= ''
                    sku = str(json_variant['sku']).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    wholesale_price = str(json_variant['wholesale_price']).strip()
                    listing_price = str(json_variant['listing_price']).strip()

                    image_filname = f'Images/{sku}.jpg'
                    if not os.path.exists(image_filname):
                        image_attachment = download_image(img_url)
                        if image_attachment:
                            with open(f'Images/{sku}.jpg', 'wb') as f: f.write(image_attachment)
                            crop_downloaded_image(f'Images/{sku}.jpg')

                    data.append([name, sku, frame_color, lens_color, listing_price, wholesale_price])
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_image(url):
    image_attachment = ''
    try:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-Encoding': 'gzip, deflate, br',
            'accept-Language': 'en-US,en;q=0.9',
            'cache-Control': 'max-age=0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'Sec-Fetch-User': '?1',
            'upgrade-insecure-requests': '1',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=20)
                # print(response.status_code)
                if response.status_code == 200:
                    # image_attachment = base64.b64encode(response.content)
                    image_attachment = response.content
                    break
                else: print(f'{response.status_code} found for downloading image')
            except: sleep(0.3)
            counter += 1
            if counter == 10: break
    except Exception as e: print(f'Exception in download_image: {str(e)}')
    finally: return image_attachment

def crop_downloaded_image(filename):
    try:
        im = Image.open(filename)
        width, height = im.size   # Get dimensions
        new_width = 1120
        new_height = 600
        if width > new_width and height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
        elif height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
    except Exception as e: print(f'Exception in crop_downloaded_image: {e}')

def saving_picture_in_excel(data: list):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Name')
    worksheet.cell(row=1, column=2, value='SKU')
    worksheet.cell(row=1, column=3, value='Frame Color')
    worksheet.cell(row=1, column=4, value='Lens Color')
    worksheet.cell(row=1, column=5, value='Listing Price')
    worksheet.cell(row=1, column=6, value='Wholesale Price')
    worksheet.cell(row=1, column=10, value="Image")

    for index, d in enumerate(data):
        new_index = index + 2

        worksheet.cell(row=new_index, column=1, value=d[0])
        worksheet.cell(row=new_index, column=2, value=d[1])
        worksheet.cell(row=new_index, column=3, value=d[2])
        worksheet.cell(row=new_index, column=4, value=d[3])
        worksheet.cell(row=new_index, column=5, value=d[4])
        worksheet.cell(row=new_index, column=6, value=d[5])

        image = f'Images/{d[1]}.jpg'
        if os.path.exists(image):
            im = Image.open(image)
            width, height = im.size
            worksheet.row_dimensions[new_index].height = height
            worksheet.add_image(Imag(image), anchor='G'+str(new_index))
            # col_letter = get_column_letter(7)
            # worksheet.column_dimensions[col_letter].width = width

    workbook.save('Rudy Project Results.xlsx')

DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    
    if os.path.exists('RudyProject Results.xlsx'): os.remove('RudyProject Results.xlsx')

    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False
    
    f = open('requirements/rudyproject.json')
    data = json.loads(f.read())
    f.close()

    store = Store()
    store.link = data['url']

    result_filename = 'requirements/RudyProject Results.json'

    if not os.path.exists('Logs'): os.makedirs('Logs')

    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')

    scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
    logs_filename = f'Logs/Logs {scrape_time}.txt'
    
    RudyProject_Scraper(DEBUG, result_filename, logs_filename).controller(store)
    
    for filename in glob.glob('Images/*'): os.remove(filename)
    print('Downloading images...\n')
    data = read_data_from_json_file(DEBUG, result_filename)
    os.remove(result_filename)

    saving_picture_in_excel(data)
except Exception as e:
    if DEBUG: print('Exception: '+str(e))
    else: pass
